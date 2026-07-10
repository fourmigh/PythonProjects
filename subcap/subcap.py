import subprocess
import sys
import urllib.request


def check_and_install_deps():
    pkg_map = {
        'opencv-python': 'cv2',
        'rapidocr-onnxruntime': 'rapidocr_onnxruntime',
        'openpyxl': 'openpyxl',
        'numpy': 'numpy',
    }
    missing = []
    for pkg, mod in pkg_map.items():
        try:
            __import__(mod)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"Installing Python packages: {' '.join(missing)}")
        subprocess.check_call([sys.executable, '-m', 'pip', 'install'] + missing)


check_and_install_deps()


import cv2
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import os
import re
import numpy as np
from rapidocr_onnxruntime import RapidOCR
from difflib import SequenceMatcher


_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_YUNET_URL = "https://github.com/opencv/opencv_zoo/raw/main/models/face_detection_yunet/face_detection_yunet_2023mar.onnx"
_face_net = None
_ocr = None


def _init_face_detector():
    global _face_net
    if _face_net is not None:
        return
    model_path = os.path.join(_SCRIPT_DIR, "face_detection_yunet_2023mar.onnx")
    if not os.path.exists(model_path):
        print("[INFO] Downloading YuNet face detector...")
        urllib.request.urlretrieve(_YUNET_URL, model_path)
    _face_net = cv2.FaceDetectorYN.create(model_path, "", (320, 320))
    print("[INFO] Face detector ready (YuNet ONNX)")


def _init_ocr():
    global _ocr
    if _ocr is not None:
        return
    print("[INFO] Initializing RapidOCR...")
    _ocr = RapidOCR(use_text_det=False, print_verbose=False)
    print("[INFO] RapidOCR ready")



_AD_CHARS = {
    '邀', '您', '观', '看', '精', '彩', '剧', '集', '请', '欣', '赏', '采',
    '鲜', '活', '营', '养', '母', '源', '自', '护',
    '时', '光', '华', '油', '修', '抗', '皱', '耐', '老',
    '长', '焦', '我', '喜', '欢', '每', '秒', '都', '生', '动', '神', '器',
    '京', '东', '新', '品', '抢', '先', '体', '验', '重', '态',
    '龙', '牡', '壮', '骨', '健', '脾', '加', '钙', '助', '高',
    '迹', '茶', '评', '审', '状',
}


def _is_subtitle_ad(text):
    chinese = re.sub(r'[^\u4e00-\u9fff]', '', text)
    if len(chinese) < 4:
        return True

    english = re.sub(r'[^a-zA-Z]', '', text)
    if len(english) > len(chinese) * 2:
        return True

    hits = sum(1 for c in chinese if c in _AD_CHARS)
    if hits >= 3 and hits / len(chinese) >= 0.5:
        return True

    return False


_WATERMARK_AD_CHARS = {'广', '告', '广 告'}


def _has_ad_watermark(frame, height, width, x1, y1, x2, y2):
    wm_roi = frame[int(height * y1):int(height * y2),
                   int(width * x1):int(width * x2)]
    if wm_roi.size == 0:
        return False
    try:
        result, _ = _ocr(wm_roi)
        if result and len(result) > 0:
            wm_text = result[0][1].strip()
            for c in _WATERMARK_AD_CHARS:
                if c in wm_text:
                    return True
    except Exception:
        pass
    return False


def ms_to_timestr(ms):
    """毫秒转 HH:MM:SS.mmm 格式"""
    seconds, ms = divmod(int(ms), 1000)
    minutes, seconds = divmod(seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}.{ms:03d}"


def sanitize_filename(text):
    text = re.sub(r'[\\/*?:"<>|]', '', text)
    text = re.sub(r'\s+', '_', text)
    return text[:50]


def score_frame_quality(frame):
    """
    对画面质量进行综合评分（0-100分）
    评分维度：清晰度、人脸质量、构图、亮度
    """
    score = 0
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    height, width = frame.shape[:2]

    # 1. 清晰度评分（拉普拉斯方差，越高越清晰）
    laplacian_var = cv2.Laplacian(gray, cv2.CV_64F).var()
    clarity_score = min(laplacian_var / 1000, 30)  # 最高30分
    score += clarity_score

    # 2. 人脸检测与质量评分
    face_score = 0
    if _face_net is not None:
        _face_net.setInputSize((width, height))
        _, faces = _face_net.detect(frame)

        valid_faces = []
        if faces is not None:
            for i in range(faces.shape[0]):
                x1, y1, w, h, confidence = faces[i, :5]
                if confidence > 0.5:
                    valid_faces.append((int(x1), int(y1), int(w), int(h)))

        if valid_faces:
            num_faces = len(valid_faces)
            face_score += min(num_faces * 8, 25)

            for (x, y, w, h) in valid_faces:
                center_x = (x + w / 2) / width
                center_y = (y + h / 2) / height
                if 0.3 < center_x < 0.7 and 0.3 < center_y < 0.7:
                    face_score += 10
                face_ratio = (w * h) / (width * height)
                if 0.01 < face_ratio < 0.15:
                    face_score += 5

                rw = min(w, width - x)
                rh = min(h, height - y)
                if rw > 20 and rh > 20:
                    fx = max(0, x)
                    fy = max(0, y)
                    face_roi = gray[fy:fy+rh, fx:fx+rw]
                    eye_roi = face_roi[:int(rh*0.4), :]
                    if eye_roi.size > 0:
                        eye_var = cv2.Laplacian(eye_roi, cv2.CV_64F).var()
                        if eye_var > 50:
                            face_score += 10

    score += min(face_score, 35)  # 人脸最高35分

    # 3. 构图评分
    center_roi = gray[int(height*0.25):int(height*0.75),
                      int(width*0.25):int(width*0.75)]
    if center_roi.size > 0:
        center_var = cv2.Laplacian(center_roi, cv2.CV_64F).var()
        if center_var > 100:
            score += 20
        elif center_var > 50:
            score += 10

    # 4. 亮度评分
    mean_brightness = np.mean(gray)
    if 80 < mean_brightness < 200:
        score += 15
    elif 50 < mean_brightness < 220:
        score += 5

    return min(score, 100)


def save_best_frame(text, frames, output_dir, ws, idx, start_ms, end_ms,
                    frame_count=0, total_frames=0):
    """
    从一组帧中选出质量最高的保存，并写入Excel
    """
    if not frames:
        return

    scored_frames = []
    for frame in frames:
        quality_score = score_frame_quality(frame)
        scored_frames.append((quality_score, frame))

    scored_frames.sort(key=lambda x: x[0], reverse=True)
    best_score, best_frame = scored_frames[0]

    safe_text = sanitize_filename(text)
    filename = f"{idx:04d}_{safe_text}.jpg"
    filepath = os.path.join(output_dir, filename)
    cv2.imwrite(filepath, best_frame, [cv2.IMWRITE_JPEG_QUALITY, 90])

    start_time = ms_to_timestr(start_ms)
    end_time = ms_to_timestr(end_ms)

    row_num = idx + 1
    ws.cell(row=row_num, column=1, value=idx)
    ws.cell(row=row_num, column=2, value=start_time)
    ws.cell(row=row_num, column=3, value=end_time)
    ws.cell(row=row_num, column=4, value=text)
    ws.cell(row=row_num, column=5, value=filename)
    ws.cell(row=row_num, column=6, value=f"{best_score:.1f}")

    try:
        img = XLImage(filepath)
        img.width = 120
        img.height = 68
        ws.add_image(img, f'G{row_num}')
        ws.row_dimensions[row_num].height = 75
    except Exception:
        pass

    pct = (frame_count / total_frames) * 100 if total_frames else 0
    print("\n[SAVED #{}] {}/{} ({:.1f}%) \"{}\"... [{}->{}] score={:.1f}".format(
        idx, frame_count, total_frames, pct, text[:30],
        start_time, end_time, best_score))


def extract_best_frames(video_path, output_dir, excel_path,
                        similarity_threshold=0.6,
                        min_text_length=4,
                        skip_frames=2):
    """
    提取字幕并保存画面质量最优的帧
    """
    if not os.path.exists(video_path):
        print("[ERROR] Video file not found: " + video_path)
        return

    os.makedirs(output_dir, exist_ok=True)

    video_dir = os.path.dirname(os.path.abspath(video_path))
    config_path = os.path.join(video_dir, "roi_config.json")
    roi_cfg = {}
    if os.path.exists(config_path):
        import json
        with open(config_path) as f:
            roi_cfg = json.load(f)

    sx1 = roi_cfg.get("subtitle_x1", 0.0)
    sx2 = roi_cfg.get("subtitle_x2", 1.0)
    sy1 = roi_cfg.get("subtitle_y1", 0.65)
    sy2 = roi_cfg.get("subtitle_y2", 0.95)
    wm_x1 = roi_cfg.get("wm_x1", 0.65)
    wm_y1 = roi_cfg.get("wm_y1", 0.88)
    wm_x2 = roi_cfg.get("wm_x2", 1.0)
    wm_y2 = roi_cfg.get("wm_y2", 1.0)

    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['序号', '开始时间', '结束时间', '台词', '截图文件名', '画面质量分', '截图预览'])
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 30

    print(f"[INFO] Starting frame detection, total frames: {total_frames}")
    print("Collecting subtitle frames...\n")

    _init_face_detector()
    _init_ocr()

    frame_count = 0
    current_text = ""
    current_frames = []
    current_start_ms = 0
    current_end_ms = 0
    idx = 1

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        frame_count += 1

        pct = (frame_count / total_frames) * 100
        print("\r[ALIVE] frame {} / {} ({:.1f}%)".format(
            frame_count, total_frames, pct), end="", flush=True)

        if frame_count % skip_frames != 0:
            continue

        timestamp_ms = cap.get(cv2.CAP_PROP_POS_MSEC)

        height, width = frame.shape[:2]

        roi = frame[int(height * sy1):int(height * sy2),
                    int(width * sx1):int(width * sx2)]

        try:
            result, _ = _ocr(roi)
            if result and len(result) > 0:
                text = result[0][1].strip()
            else:
                continue
        except Exception:
            continue

        if _is_subtitle_ad(text):
            continue

        if _has_ad_watermark(frame, height, width, wm_x1, wm_y1, wm_x2, wm_y2):
            continue

        clean_text = re.sub(r'[^a-zA-Z\u4e00-\u9fa5]', '', text)
        if len(clean_text) < min_text_length:
            if not current_text:
                continue
            else:
                save_best_frame(current_text, current_frames, output_dir, ws,
                               idx, current_start_ms, current_end_ms,
                         frame_count=frame_count, total_frames=total_frames)
                idx += 1
                current_text = ""
                current_frames = []
                continue

        if current_text:
            current_clean = re.sub(r'[^a-zA-Z\u4e00-\u9fa5]', '', current_text)
            similarity = SequenceMatcher(None, clean_text, current_clean).ratio()
            if similarity > similarity_threshold:
                current_frames.append(frame)
                current_end_ms = timestamp_ms
                continue
            else:
                save_best_frame(current_text, current_frames, output_dir, ws,
                               idx, current_start_ms, current_end_ms,
                         frame_count=frame_count, total_frames=total_frames)
                idx += 1
                current_text = text
                current_frames = [frame]
                current_start_ms = timestamp_ms
                current_end_ms = timestamp_ms
        else:
            current_text = text
            current_frames = [frame]
            current_start_ms = timestamp_ms
            current_end_ms = timestamp_ms

    if current_text and current_frames:
        save_best_frame(current_text, current_frames, output_dir, ws,
                       idx, current_start_ms, current_end_ms,
                         frame_count=frame_count, total_frames=total_frames)
        idx += 1

    cap.release()
    wb.save(excel_path)
    print(f"\n[DONE] Extracted {idx - 1} subtitles")
    print(f"       Screenshots: {output_dir}")
    print(f"       Excel: {excel_path}")


if __name__ == "__main__":
    video_dir = "videos"
    if not os.path.isdir(video_dir):
        print("[ERROR] Video directory not found: " + video_dir)
        sys.exit(1)

    extensions = ('.mp4', '.mkv', '.avi', '.mov', '.flv', '.webm', '.wmv')
    video_files = [f for f in sorted(os.listdir(video_dir))
                   if f.lower().endswith(extensions)]

    if not video_files:
        print("[ERROR] No video files found in: " + video_dir)
        sys.exit(1)

    print("Found {} video(s) to process\n".format(len(video_files)))

    for filename in video_files:
        video_path = os.path.join(video_dir, filename)
        name = os.path.splitext(filename)[0]
        output_dir = os.path.join(video_dir, name + "_screenshots")
        excel_path = os.path.join(video_dir, name + ".xlsx")

        print("=" * 60)
        print("Processing: " + filename)
        print("=" * 60)

        extract_best_frames(
            video_path=video_path,
            output_dir=output_dir,
            excel_path=excel_path,
            similarity_threshold=0.7,
            min_text_length=4,
            skip_frames=2
        )
